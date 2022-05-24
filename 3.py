#coding=utf-8
from csv import excel
from ntpath import join
from openpyxl import load_workbook
list_ip = []
list_port = []
list_portco2 = []
list_portco1 = []
# print(type(name))
name = input("输入时间")
wb_total = load_workbook("./FTTH光衰整治通报%s.xlsx"%name)
list_sheet=wb_total.sheetnames
wb_worksheet = wb_total[list_sheet[1]]
file1 = open("count.txt",'r')
list1 = file1.readlines()
list2 = []
print(list1)
for i in range(0,len(list1),2):
    list2.append(list1[i])
for i in range(len(list2)):
    list2[i] = list2[i].rstrip("\n")
print(list2)
for i in range(len(list2)):
    s = str(i+2)
    wb_worksheet["A"+s].value = list2[i]
wb_total.save("改FTTH光衰整治通报%s.xlsx"%name)
