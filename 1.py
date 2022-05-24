#coding=utf-8
from csv import excel
from ntpath import join
from openpyxl import load_workbook
name = input("输入时间")
list_ip = []
list_port = []
list_portco2 = []
list_portco1 = []
list_dt = []
# print(type(name))
wb_total = load_workbook("./FTTH光衰整治通报%s.xlsx"%name)
list_sheet=wb_total.sheetnames
print(list_sheet[1])
wb_worksheet = wb_total[list_sheet[1]]
ip_dizhi = wb_worksheet["D"]
port_dk = wb_worksheet["E"]
dtype = wb_worksheet["F"]
# print(ip_dizhi)
for i in range(len(ip_dizhi)):
    list_ip.append(ip_dizhi[i].value)
for i in range(len(port_dk)):
    list_port.append(port_dk[i].value)    
for i in range(len(dtype)):
    list_dt.append(dtype[i].value)   
# print(list_ip)
file = open("ip.txt", 'w').close() 
iptxt = open("ip.txt","w+")
for i in range(1,len(list_ip)):
    iptxt.write(list_ip[i]+'\n')
iptxt.close()
file = open("port1.txt", 'w').close() 
porttxt1 = open("port1.txt",'w+')
file = open("port2.txt", 'w').close() 
porttxt2 = open("port2.txt",'w+')
for i in range(1,len(list_port)):
    list_process1 = list_port[i].split("-")
    porttxt1.write(list_process1[-2]+'\n')
    porttxt2.write(list_process1[-1]+'\n')
porttxt1.close()
porttxt2.close()
file = open("dtype.txt", 'w').close() 
detypef = open("dtype.txt","w+")
for i in range(1,len(list_dt)):
    list_strdt1 = list_dt[i].split("-")
    list_strdt2 = list_strdt1[-1].split("/")
    strdt3 = list_strdt2[0].replace("T","")
    strdt3 = strdt3.replace("MA","")
    strdt3 = strdt3.replace("HuaWei","")
    strdt3 = strdt3.replace("C","")
    detypef.write(strdt3+'\n')
detypef.close()
